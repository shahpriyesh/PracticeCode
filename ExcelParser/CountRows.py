import xlrd
from openpyxl import *

def countEqualEntries(excel_file, sheet_name):
    wb = xlrd.open_workbook(excel_file)
    sheet = wb.sheet_by_name(sheet_name)
    counter = {}

    for i in range(1, sheet.nrows):
        row = sheet.row_values(i)
        if row[0] in counter:
            counter[row[0]][0] += 1
            if isinstance(row[-2], str):
                counter[row[0]][3].append(row[-2])
        else:
            # List contains (Count, Question string, difficulty level, list of company names)
            if 'Easy' in row:
                counter[row[0]] = [1, row[1], 'Easy', []]
            elif 'Medium' in row:
                counter[row[0]] = [1, row[1], 'Medium', []]
            else:
                counter[row[0]] = [1, row[1], 'Hard', []]

            if isinstance(row[-2], str):
                counter[row[0]][3].append(row[-2])

    wb.release_resources()
    del wb
    return counter


def writeTotalEntries(excel_file, sheet_name, counter):
    wb = load_workbook(excel_file)
    sheet = wb[sheet_name]
    unique = set()

    for i in range(1, sheet.max_row):
        c1 = sheet.cell(row=i, column=1)
        c9 = sheet.cell(row=i, column=9)
        c10 = sheet.cell(row=i, column=10)
        c11 = sheet.cell(row=i, column=11)
        # writing values to cells
        if c1.value and c1.value not in unique:
            mapping = counter[c1.value]
            c9.value = mapping[0]
            c10.value = mapping[2]
            c11.value = ', '.join(mapping[3])
            unique.add(c1.value)

    wb.save(excel_file)
    wb.close()


def reformatEntries(excel_file, sheet_name):
    unique = set()

    wb = load_workbook(excel_file)
    sheet = wb[sheet_name]

    for i in range(1, sheet.max_row):
        c1 = sheet.cell(row=i, column=1)
        if c1.value in unique:
            sheet.delete_rows(i, 1)
        unique.add(c1.value)

    wb.save(excel_file)
    wb.close()



filename = "/Users/pshah/Downloads/Leetcode_FAQ.xlsx"
sheetname = "FAQs"
counter = countEqualEntries(filename, sheetname)
writeTotalEntries(filename, sheetname, counter)
# reformatEntries(filename, sheetname)